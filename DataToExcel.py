from openpyxl import Workbook
import os
import shutil
# images_info = [(8, 12, 395, 69), (403, 12, 439, 33), (842, 12, 135, 69), (403, 45, 151, 36), (554, 45, 165, 36), (719, 45, 124, 36), (8, 80, 395, 47), (403, 80, 151, 47), (554, 80, 165, 47), (719, 80, 123, 47), (842, 80, 135, 47), (8, 127, 395, 47), (403, 127, 151, 47), (554, 127, 165, 47), (719, 127, 123, 47), (842, 127, 135, 47), (8, 174, 395, 46), (403, 174, 151, 46), (554, 174, 165, 46), (719, 174, 123, 46), (842, 174, 135, 46), (8, 220, 395, 45), (403, 220, 151, 45), (554, 220, 165, 45), (719, 220, 123, 45), (842, 220, 135, 45), (8, 265, 395, 42), (403, 265, 439, 42), (842, 265, 135, 42), (8, 306, 395, 42), (403, 306, 439, 42), (842, 306, 135, 42), (8, 348, 395, 40), (403, 348, 439, 40), (842, 348, 135, 40), (8, 387, 395, 42), (403, 387, 439, 42), (842, 387, 135, 42)]
# txts = ['Description', 'Specification', 'Tolerance', 'L1', 'L2', 'L3', 'Min.Pitch', '8', '26', None, None, 'FinishedBottom', '25', '26', '25', 'L1:+/-3', 'FinishedBottom', '8', '26', '一', 'MAX', 'Recessed', '8', None, None, None, 'Min.Trace', '26', None, 'FinishedTrace', '7', '+/-6', 'FinishedTrace', '26', '+/-6', 'Min.BumpPitch', '30', None]

def data_to_excel(dir,images_info, txts):
    # 檢查目錄
    excel_dir = 'SavedExcel'
    
    # 如果資料夾存在，則遞迴地刪除它及其底下的所有文件
    if not os.path.exists(excel_dir):
        # 創建資料夾
        os.makedirs(excel_dir)
    
    txts_filtered = ['None' if item is None else item for item in txts]
    grouped_txts_y = {}
    grouped_txts_x = {}
    # 將 txts 中的字串依照 images_info 中每個元組的第二個值進行分組

    # 建立y軸字典
    for i, item in enumerate(images_info):
        key = item[1]  # 使用每個元組的第二個值作為鍵
        if key in grouped_txts_y:
            grouped_txts_y[key].append(txts_filtered[i])
        else:
            grouped_txts_y[key] = [txts_filtered[i]]
    sorted_grouped_txts_y = dict(sorted(grouped_txts_y.items()))

    # 建立x軸字典
    for i, item in enumerate(images_info):
        key = item[0]  # 使用每個元組的第二個值作為鍵
        if key in grouped_txts_x:
            grouped_txts_x[key].append(txts_filtered[i])
        else:
            grouped_txts_x[key] = [txts_filtered[i]]
    sorted_grouped_txts_x = dict(sorted(grouped_txts_x.items()))

    # 找出x軸個數(建立座標系)
    # x_max_key = max(sorted_grouped_txts_x, key=lambda x: len(sorted_grouped_txts_x[x]))
    # x_max_value_count = len(sorted_grouped_txts_x[x_max_key])

    # 找出y軸個數(建立座標系)
    # y_max_key = max(sorted_grouped_txts_y, key=lambda y: len(sorted_grouped_txts_y[y]))
    # y_max_value_count = len(sorted_grouped_txts_y[y_max_key])


    # print(sorted_grouped_txts_x,x_max_value_count)
    # print(sorted_grouped_txts_y,y_max_value_count)

    # 將每個字的x,y與字合併
    # combined_data = [(item[0], item[1], txt) for item, txt in zip(images_info, txts_filtered)]

    # 將每個字的x,y,w,h與字合併
    combined_data = [(item[0], item[1], item[2], item[3], txt) for item, txt in zip(images_info, txts_filtered)]
    # print('combined_data', combined_data)


    # 建立座標系
    x_values = list(sorted_grouped_txts_x.keys())
    y_values = list(sorted_grouped_txts_y.keys())
    coordinates_text = [['N/A'] * len(x_values) for _ in range(len(y_values))]

    pic_coordinates = [[None] * len(x_values) for _ in range(len(y_values))]
    
    # 填充座標系
    # for x, y, w, h, data in combined_data:
    #     x_index = x_values.index(x)
    #     y_index = y_values.index(y)
    #     coordinates[y_index][x_index] = data
    
    
    for index, data in enumerate(combined_data):
        
        x, y, w, h, text = data  # 直接解包元组
        
        x_index = x_values.index(x)
        y_index = y_values.index(y)
        coordinates_text[y_index][x_index] = text
        pic_coordinates[y_index][x_index] = (x,y,w,h)
            
    wb = Workbook()
    ws = wb.active
    for row in coordinates_text:
        ws.append(row)

    # for row in pic_coordinates:
    #     print(row)

    #橫向合併
    for index, sublist in enumerate(pic_coordinates):
        first, second, distance = calculate_horizontal_gap_distance(sublist)
        if first is not None and second is not None:  # Check if first and second are not None
            if sublist[first] is not None and sublist[second] is not None:
                first_x, first_y, first_w, first_h = sublist[first]
                second_x, second_y, second_w, second_h = sublist[second]
                if distance > 0 and first_x + first_w >= second_x - 5 and first_x + first_w <= second_x + 5:
                    ws.merge_cells(start_row=index+1, start_column=first+1, end_row=index+1, end_column=second)
                    sublist[first+1:second] = [(-1, -1, -1, -1)] * (second - (first + 1))
    
    #縱向合併
    transported_pic_coordinates = [list(i) for i in zip(*pic_coordinates)] #轉置
    for index, sublist in enumerate(transported_pic_coordinates):
        first, second, distance = calculate_vertical_gap_distance(sublist)
        if first is not None and second is not None:  # Check if first and second are not None
            if sublist[first] is not None and sublist[second] is not None:
                first_x, first_y, first_w, first_h = sublist[first]
                second_x, second_y, second_w, second_h = sublist[second]
                if distance > 0 and first_y + first_h >= second_y - 5 and first_y + first_h <= second_y + 5:
                    ws.merge_cells(start_row=first+1, start_column=index+1, end_row=second, end_column=index+1)
                    sublist[first+1:second] = [(-1, -1, -1, -1)] * (second - (first + 1))
    
                    
    # for row in transported_pic_coordinates:
    #     print(row)
    file_name = dir.split('\\')
    save_dir = os.path.join(excel_dir,file_name[-1])
    save_dir = f'{save_dir}_SavedText.xlsx'
    print(save_dir)
    # 保存工作簿
    wb.save(save_dir)



def calculate_horizontal_gap_distance(sublist):
    first_non_none_index = None
    second_non_none_index = None
    distance = 0
    
    # 找到第一個非None值索引
    for i, item in enumerate(sublist[:-1]):
        if item is not None and sublist[i+1] is None:
            first_non_none_index = i
            break

    # 找到第二個非None值索引
    if first_non_none_index is not None:
        for i, item in enumerate(sublist[first_non_none_index+1:], start=first_non_none_index+1):
            if item is not None:
                second_non_none_index = i
                break

    # 計算距離
    if first_non_none_index is not None and second_non_none_index is not None:
        distance = second_non_none_index - first_non_none_index
    return first_non_none_index, second_non_none_index, distance



def calculate_vertical_gap_distance(sublist):
    first_non_none_index = None
    second_non_none_index = None
    distance = 0
    
    # 找到第一個非None值索引
    for i, item in enumerate(sublist[:-1]):
        if item is not None and sublist[i+1] is None:
            first_non_none_index = i
            break

    # 找到第二個非None值索引
    if first_non_none_index is not None:
        for i, item in enumerate(sublist[first_non_none_index+1:], start=first_non_none_index+1):
            if item is not None:
                second_non_none_index = i
                break

    # 計算距離
    if first_non_none_index is not None and second_non_none_index is not None:
        distance = second_non_none_index - first_non_none_index
    return first_non_none_index, second_non_none_index, distance

# data_to_excel(images_info, txts)